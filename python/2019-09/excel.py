import openpyxl
import json
from urllib.request import urlopen

url = "https://jsonplaceholder.typicode.com/todos"

content_str = urlopen(url).read().decode("utf-8")
todos = json.loads(content_str)
print(todos[0])

wb = openpyxl.Workbook()

# print(wb.worksheets[0])

wb.create_sheet("todos")
todos_sheet = wb["todos"]

for i, todo in enumerate(todos, 1):
    print(todo["title"])
    todos_sheet.cell(i, 1).value = todo["id"]
    todos_sheet.cell(i, 2).value = todo["userId"]
    todos_sheet.cell(i, 3).value = todo["title"]
    todos_sheet.cell(i, 4).value = todo["completed"]

wb.save("wb.xlsx")
